import yt_dlp
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os


def download_video(url):
    ydl_opts = {'format': 'mp4'}
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        info_dict = ydl.extract_info(url, download=False)
        video_title = info_dict.get('title', None)
        video_filename = ydl.prepare_filename(info_dict)
        ydl.download([url])
        return video_title, video_filename


def save_to_excel(video_title, video_filename):
    workbook = Workbook()
    sheet = workbook.active

    # Write column headers
    sheet['A1'] = 'File Name'
    sheet['B1'] = 'File Size'
    sheet['C1'] = 'File Date'
    sheet['D1'] = 'Raw Bytes'

    # Retrieve file metadata
    file_size = os.path.getsize(video_filename)
    file_date = os.path.getctime(video_filename)

    # Write metadata to Excel sheet
    sheet['A2'] = video_filename
    sheet['B2'] = file_size
    sheet['C2'] = file_date

    # Save the raw bytes to Excel sheet
    with open(video_filename, 'rb') as f:
        video_bytes = f.read()
        sheet.append([None, None, None, video_bytes])

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    workbook.save('video_metadata.xlsx')


# Example usage
video_url = 'https://www.youtube.com/watch?v=29vYgKKtLjA'
title, filename = download_video(video_url)
save_to_excel(title, filename)
